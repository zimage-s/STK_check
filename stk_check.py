#!/usr/bin/env python3
"""
Kontrola STK vozidel z XLSX souboru.

Primární zdroj: API Ministerstva dopravy (dataovozidlech.cz)
Fallback:       kontrolatachaku.cz (pro VINy, které API nezná)

Použití:
  python3 stk_check.py stahni              # stáhne STK data → uloží do stk_data.json
  python3 stk_check.py xlsx                # z stk_data.json vygeneruje STK_vysledky.xlsx
  python3 stk_check.py stahni xlsx         # obojí najednou
  python3 stk_check.py stahni --vin VIN    # stáhne jen jedno auto podle VIN (doplní do JSONu)
"""

import sys
import requests
import re
import time
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from bs4 import BeautifulSoup
from datetime import datetime, date

INPUT_FILE = "Přehled vozidel Ing. Pavel Zima.xlsx"
OUTPUT_FILE = "STK_vysledky.xlsx"
DATA_FILE = "stk_data.json"

MD_API_URL = "https://api.dataovozidlech.cz/api/vehicletechnicaldata/v2"
CONFIG_FILE = "config.json"

def _load_api_key():
    try:
        with open(CONFIG_FILE, encoding="utf-8") as f:
            return json.load(f).get("md_api_key", "")
    except FileNotFoundError:
        return ""

MD_API_KEY = _load_api_key()

UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"


# ---------------------------------------------------------------------------
#  Načtení aut z XLSX
# ---------------------------------------------------------------------------

def load_cars():
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb["List1"]
    cars = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        num = row[0].value
        if num is None:
            continue
        cars.append({
            "num": num,
            "brand": row[7].value or "",
            "model": row[8].value or "",
            "rz":    row[10].value or "",
            "vin":   row[11].value or "",
            "year":  row[13].value if row[13].value else "",
        })
    return cars


# ---------------------------------------------------------------------------
#  Stažení STK dat — oficiální API MD ČR
# ---------------------------------------------------------------------------

def fetch_stk_api(vin):
    """Dotaz na API Ministerstva dopravy (dataovozidlech.cz)."""
    r = requests.get(
        MD_API_URL,
        params={"vin": vin},
        headers={"api_key": MD_API_KEY},
        timeout=15,
    )
    if r.status_code != 200:
        return None, f"API HTTP {r.status_code}"

    resp = r.json()
    d = resp.get("Data")
    if not d:
        status = resp.get("Status", "?")
        return None, f"API nenalezeno (status {status})"

    znacka = d.get("TovarniZnacka", "")
    model = d.get("ObchodniOznaceni", "")
    typ = d.get("Typ", "")
    web_znacka = " ".join(filter(None, [znacka, model, typ])).strip()
    web_druh = " ".join(filter(None, [
        d.get("VozidloDruh", ""),
        d.get("VozidloDruh2", ""),
    ])).strip()
    kategorie = d.get("Kategorie", "")
    if kategorie:
        web_druh = f"{web_druh} ({kategorie})".strip()

    pristi_stk = None
    zbyva_dni = None
    po_lhute_dni = None
    stk_raw = d.get("PravidelnaTechnickaProhlidkaDo")
    if stk_raw:
        try:
            stk_date = datetime.fromisoformat(stk_raw).date()
            pristi_stk = stk_date.strftime("%d.%m.%Y")
            delta = (stk_date - date.today()).days
            if delta < 0:
                po_lhute_dni = str(abs(delta))
            else:
                zbyva_dni = f"{delta} dní"
        except (ValueError, TypeError):
            pass

    result = {
        "records": [],
        "pristi_stk": pristi_stk,
        "zbyva_dni": zbyva_dni,
        "po_lhute_dni": po_lhute_dni,
        "web_znacka": web_znacka,
        "web_druh": web_druh,
        "zdroj": "api.dataovozidlech.cz",
        "api_data": {
            "znacka": znacka,
            "model": model,
            "typ": typ,
            "druh": d.get("VozidloDruh", ""),
            "druh2": d.get("VozidloDruh2", ""),
            "kategorie": kategorie,
            "status": d.get("StatusNazev", ""),
            "barva": d.get("VozidloKaroserieBarva", ""),
            "cislo_tp": d.get("CisloTp", ""),
            "datum_prvni_registrace": d.get("DatumPrvniRegistrace", ""),
            "datum_registrace_cr": d.get("DatumPrvniRegistraceVCr", ""),
            "pocet_vlastniku": d.get("PocetVlastniku", ""),
        },
    }
    return result, None


# ---------------------------------------------------------------------------
#  Stažení STK dat — fallback přes kontrolatachaku.cz
# ---------------------------------------------------------------------------

def fetch_stk_web(vin):
    """Scraping z kontrolatachaku.cz (pro VINy, které API nezná)."""
    session = requests.Session()
    session.headers.update({"User-Agent": UA})

    r = session.get("https://www.kontrolatachaku.cz")
    csrf_match = re.search(r'name="csrf_token"\s+value="([^"]+)"', r.text)
    if not csrf_match:
        return None, "CSRF token nenalezen"

    r = session.post(
        "https://www.kontrolatachaku.cz",
        data={"vin": vin, "csrf_token": csrf_match.group(1), "hp_field": ""},
    )
    if r.status_code != 200:
        return None, f"HTTP {r.status_code}"

    soup = BeautifulSoup(r.text, "html.parser")
    table = soup.find("table", id="resultsTable")
    if not table or len(table.find_all("tr")) < 2:
        return None, "Žádné záznamy STK"

    records = []
    for row in table.find_all("tr")[1:]:
        cells = row.find_all("td")
        if len(cells) >= 5:
            records.append({
                "datum":    cells[0].get_text(strip=True),
                "druh":     cells[1].get_text(strip=True),
                "km":       cells[2].get_text(strip=True),
                "pristi":   cells[3].get_text(strip=True),
                "vysledek": cells[4].get_text(strip=True),
            })

    stk_only = [r for r in records if "videnc" not in r["druh"].lower() and "videnčn" not in r["druh"].lower()]
    latest = stk_only[0] if stk_only else None

    pristi_match  = re.search(r"Příští prohlídka je naplánována na ([^,]+), do které zbývá přibližně ([^.]+)\.", r.text)
    po_lhute_match = re.search(r"po lhůtě již (\d+)", r.text)
    if not po_lhute_match and latest:
        m = re.search(r"(\d+)\s*dn[íi]\s*po\s*lhůtě", latest["pristi"])
        if m:
            po_lhute_match = m
    if pristi_match and po_lhute_match:
        po_lhute_match = None

    pristi_stk = None
    if pristi_match:
        pristi_stk = pristi_match.group(1)
    elif latest and latest["pristi"] != "—":
        m = re.search(r"(\d{2}\.\d{2}\.\d{4})", latest["pristi"])
        if m:
            pristi_stk = m.group(1)

    web_znacka = web_druh = ""
    detail_tables = soup.find_all("table")
    for dt in detail_tables:
        if dt.get("id") == "resultsTable":
            continue
        for tr in dt.find_all("tr"):
            cells = [c.get_text(strip=True) for c in tr.find_all(["td", "th"])]
            if len(cells) == 2:
                key = cells[0].lower()
                if "značka" in key and "model" in key:
                    web_znacka = re.sub(r"\s+", " ", cells[1]).strip()
                elif "druh" in key and "kategori" in key:
                    web_druh = re.sub(r"\s+", " ", cells[1]).strip()
        if web_znacka:
            break

    result = {
        "records":      records,
        "pristi_stk":   pristi_stk,
        "zbyva_dni":    pristi_match.group(2) if pristi_match else None,
        "po_lhute_dni": po_lhute_match.group(1) if po_lhute_match else None,
        "web_znacka":   web_znacka,
        "web_druh":     web_druh,
        "zdroj":        "kontrolatachaku.cz",
    }
    if latest:
        result["posledni_datum"]    = latest["datum"]
        result["posledni_km"]       = latest["km"]
        result["posledni_vysledek"] = latest["vysledek"]
        result["posledni_druh"]     = latest["druh"]
    elif records:
        result["posledni_datum"]    = records[0]["datum"]
        result["posledni_km"]       = records[0]["km"]
        result["posledni_vysledek"] = records[0]["vysledek"]
        result["posledni_druh"]     = records[0]["druh"]

    return result, None


def fetch_stk(vin):
    """Zkusí API MD, při neúspěchu fallback na kontrolatachaku.cz."""
    data, err = fetch_stk_api(vin)
    if data:
        return data, None
    data_web, err_web = fetch_stk_web(vin)
    if data_web:
        return data_web, None
    return None, err_web or err


def check_vehicle_match(car, stk_data):
    """Compare vehicle brand from XLSX vs web. Returns mismatch description or None."""
    web = stk_data.get("web_znacka", "").upper()
    if not web:
        return None
    xlsx_brand = car["brand"].upper().strip()
    # Normalize common variants
    norm = {"ŠKODA": "SKODA", "ŠKODA": "SKODA"}
    wb = norm.get(xlsx_brand, xlsx_brand)
    # Check if XLSX brand appears anywhere in web response
    web_norm = web.replace("Š", "S").replace("š", "s").upper()
    wb_norm = wb.replace("Š", "S").replace("š", "s")
    if wb_norm and wb_norm in web_norm:
        return None
    web_druh = stk_data.get("web_druh", "")
    return f"v tabulce '{car['brand']} {car['model']}', web vrací '{web}' ({web_druh})"


def cmd_stahni(only_vin=None):
    cars = load_cars()
    print(f"Načteno {len(cars)} vozidel z {INPUT_FILE}")

    try:
        with open(DATA_FILE, encoding="utf-8") as f:
            all_data = json.load(f)
    except FileNotFoundError:
        all_data = {}

    if only_vin:
        cars = [c for c in cars if c["vin"] == only_vin]
        if not cars:
            print(f"VIN {only_vin} nenalezen v tabulce!")
            return
        print(f"Stahuji jen VIN: {only_vin}")

    ok = err = 0
    last_was_web = False
    for i, car in enumerate(cars):
        label = f"{car['brand']} {car['model']}".strip()
        print(f"[{i+1:2}/{len(cars)}] {label:<30} RZ: {car['rz']:<10} VIN: {car['vin']:<22} ", end="", flush=True)

        if last_was_web:
            time.sleep(3)

        data, error = fetch_stk(car["vin"])
        last_was_web = data is not None and data.get("zdroj") == "kontrolatachaku.cz"
        if data is None and error:
            last_was_web = True

        if error:
            print(f"⚠ {error}")
            all_data[car["vin"]] = {"error": error}
            err += 1
        else:
            po = data.get("po_lhute_dni")
            zdroj = data.get("zdroj", "?")
            src = "API" if "api" in zdroj else "WEB"
            n = len(data.get("records", []))
            if po:
                print(f"⛔ PO LHŮTĚ {po} dní [{src}]")
            else:
                print(f"✅ příští: {data.get('pristi_stk','?')} [{src}]")

            mismatch = check_vehicle_match(car, data)
            if mismatch:
                print(f"     ⚠️  NESOUHLASÍ: {mismatch}")

            all_data[car["vin"]] = data
            ok += 1

    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*60}")
    print(f"Hotovo: {ok} s daty, {err} bez dat")
    print(f"Uloženo do {DATA_FILE}")


# ---------------------------------------------------------------------------
#  Generování XLSX
# ---------------------------------------------------------------------------

def cmd_xlsx():
    cars = load_cars()
    try:
        with open(DATA_FILE, encoding="utf-8") as f:
            all_data = json.load(f)
    except FileNotFoundError:
        print(f"Soubor {DATA_FILE} nenalezen. Nejdřív spusťte: python3 stk_check.py stahni")
        return

    today = date.today()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "STK Kontrola"

    hfont = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="475C7F", end_color="475C7F", fill_type="solid")
    ok_f  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    wrn_f = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    bad_f = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gry_f = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    brd   = Border(left=Side(style="thin"), right=Side(style="thin"),
                   top=Side(style="thin"), bottom=Side(style="thin"))
    center = Alignment(horizontal="center", wrap_text=True)

    headers = [
        "č.", "Značka", "Model", "RZ", "VIN", "Rok",
        "Poslední STK", "Druh", "Výsledek", "Stav km",
        "Příští STK", "Zbývá / po lhůtě", "Stav",
        "Počet záznamů", "Poznámka",
    ]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hfont, hfill, center, brd

    for i, car in enumerate(cars):
        r = i + 2
        raw = all_data.get(car["vin"])
        res = raw if raw and "error" not in (raw or {}) else None
        err_msg = raw.get("error") if isinstance(raw, dict) and "error" in raw else None

        for c, v in enumerate([car["num"], car["brand"], car["model"],
                               car["rz"], car["vin"], car["year"]], 1):
            ws.cell(row=r, column=c, value=v).border = brd

        mismatch = check_vehicle_match(car, res) if res else None

        if res is None or mismatch:
            note = mismatch or err_msg or "Žádné záznamy v DB"
            nc = ws.cell(row=r, column=15, value=note)
            nc.border = brd
            nc.fill = bad_f if mismatch else gry_f
            if mismatch:
                nc.font = Font(bold=True, color="CC0000")
            for c in range(7, 15):
                ws.cell(row=r, column=c, value="—").border = brd
            continue

        is_api = res.get("zdroj", "") == "api.dataovozidlech.cz"
        ws.cell(row=r, column=7,  value=res.get("posledni_datum", "—")).border = brd
        ws.cell(row=r, column=8,  value=res.get("posledni_druh", "—")).border = brd
        ws.cell(row=r, column=9,  value=res.get("posledni_vysledek", "—")).border = brd
        ws.cell(row=r, column=10, value=res.get("posledni_km", "—")).border = brd
        ws.cell(row=r, column=11, value=res.get("pristi_stk", "—")).border = brd

        po_lhute = res.get("po_lhute_dni")
        zbyva    = res.get("zbyva_dni")

        p = res.get("pristi_stk")
        if p and not po_lhute and not zbyva:
            try:
                d = datetime.strptime(p, "%d.%m.%Y").date()
                delta = (d - today).days
                if delta < 0:
                    po_lhute = str(abs(delta))
                else:
                    zbyva = f"{delta} dní"
            except ValueError:
                pass

        if po_lhute:
            zc = ws.cell(row=r, column=12, value=f"{po_lhute} dní po lhůtě")
            sc = ws.cell(row=r, column=13, value="PO LHŮTĚ")
            sc.fill = zc.fill = bad_f
        elif zbyva:
            zc = ws.cell(row=r, column=12, value=zbyva)
            try:
                days = int(re.search(r"\d+", zbyva).group())
            except (ValueError, AttributeError):
                days = 999
            if days <= 30:
                sc = ws.cell(row=r, column=13, value="BRZY VYPRŠÍ")
                sc.fill = zc.fill = wrn_f
            elif days <= 90:
                sc = ws.cell(row=r, column=13, value="BLÍŽÍ SE")
                sc.fill = zc.fill = wrn_f
            else:
                sc = ws.cell(row=r, column=13, value="OK")
                sc.fill = ok_f
        else:
            zc = ws.cell(row=r, column=12, value="—")
            sc = ws.cell(row=r, column=13, value="BEZ STK")
            sc.fill = gry_f

        zc.border = sc.border = brd

        n = len(res.get("records", []))
        ws.cell(row=r, column=14, value=n if n else "—").border = brd
        zdroj = res.get("zdroj", "")
        note = "API MD" if "api" in zdroj else "web"
        api_d = res.get("api_data")
        if api_d and api_d.get("status"):
            note += f" | {api_d['status']}"
        ws.cell(row=r, column=15, value=note).border = brd

    widths = [5, 12, 20, 12, 22, 8, 14, 18, 12, 14, 14, 20, 14, 10, 22]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws.auto_filter.ref = f"A1:O{len(cars)+1}"
    ws.freeze_panes = "A2"

    # --- List 2: Historie STK ---
    ws2 = wb.create_sheet("Historie STK")
    hh = ["č.", "Značka", "Model", "RZ", "VIN",
          "Datum STK (platí od)", "Druh", "Stav km", "Příští STK (platí do)", "Výsledek"]
    for c, h in enumerate(hh, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font, cell.fill, cell.alignment, cell.border = hfont, hfill, center, brd

    hr = 2
    for car in cars:
        raw = all_data.get(car["vin"])
        if not raw or "error" in (raw or {}):
            continue
        if check_vehicle_match(car, raw):
            continue
        for rec in raw.get("records", []):
            for c, v in enumerate([car["num"], car["brand"], car["model"],
                                   car["rz"], car["vin"]], 1):
                ws2.cell(row=hr, column=c, value=v).border = brd
            ws2.cell(row=hr, column=6, value=rec["datum"]).border = brd
            ws2.cell(row=hr, column=7, value=rec["druh"]).border = brd
            ws2.cell(row=hr, column=8, value=rec["km"]).border = brd
            ws2.cell(row=hr, column=9, value=rec["pristi"]).border = brd
            vc = ws2.cell(row=hr, column=10, value=rec["vysledek"])
            vc.border = brd
            if "nevyhovuje" in rec["vysledek"].lower():
                vc.fill = bad_f
            hr += 1

    hw = [5, 12, 20, 12, 22, 18, 18, 14, 30, 14]
    for c, w in enumerate(hw, 1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws2.auto_filter.ref = f"A1:J{hr-1}"
    ws2.freeze_panes = "A2"

    wb.save(OUTPUT_FILE)
    print(f"Uloženo: {OUTPUT_FILE}")
    print(f"  - STK Kontrola:  {len(cars)} vozidel")
    print(f"  - Historie STK:  {hr-2} záznamů")


# ---------------------------------------------------------------------------
#  CLI
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    if not args:
        print(__doc__)
        return

    only_vin = None
    if "--vin" in args:
        idx = args.index("--vin")
        if idx + 1 < len(args):
            only_vin = args[idx + 1]
            args = [a for a in args if a != "--vin" and a != only_vin]

    for cmd in args:
        if cmd == "stahni":
            cmd_stahni(only_vin)
        elif cmd == "xlsx":
            cmd_xlsx()
        else:
            print(f"Neznámý příkaz: {cmd}")
            print(__doc__)


if __name__ == "__main__":
    main()
